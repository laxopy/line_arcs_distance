from openpyxl import load_workbook
workbook = load_workbook(filename="coordenadas.xlsx")
workbook.sheetnames


sheet = workbook.active
sheet


for value in sheet.iter_rows(min_row=2, min_col=4,max_col=5, values_only=True):
    X1 = value
    print(X1)




